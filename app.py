from flask import Flask, render_template, request, session, redirect, url_for, make_response
from datetime import datetime, timedelta
from openpyxl import workbook, load_workbook
import pandas as pd
import os, time
from dateutil.relativedelta import relativedelta

app = Flask(__name__)
app.secret_key = 'Golf3001' 

file_path = os.path.join(app.static_folder, 'PG_2024.xlsx')  

if not os.path.exists(file_path):
    raise FileNotFoundError(f"Excel file not found at{file_path}")

try:
    workbook = load_workbook(file_path)  
    sheet = workbook["ages"]
except Exception as e:
    raise Exception(f"Error loading excel_file: {str(e)}")

age_list = []
for col in sheet.iter_cols(min_row=2, min_col=3, max_col=3, max_row=13):
    for cell in col:
        age_list.append(cell.value)
            
name_list = []
for col in sheet.iter_cols(min_row=2, min_col=0, max_col=1, max_row=13):
    for cell in col:
        name_list.append(cell.value)

members_dict_datetime = {}
members_dict_string = {}
for i, member in enumerate(name_list):
    members_dict_datetime[member] = age_list[i]
    members_dict_string[member] = age_list[i].strftime('%d-%m-%Y')        

def calculate_average_age():  
    ages = []
    today = datetime.now()
    for birthday in members_dict_string.values():
        date_format = '%d-%m-%Y'
        birthday = datetime.strptime(birthday, date_format)
        age = (today - birthday).days / 365.25
        ages.append(age)
        average_age = sum(ages) / len(ages)

    return '%.2f' %average_age

def calculate_indiv_age(birthdate):
    today = datetime.today()
    years = today.year - birthdate.year
    months = today.month - birthdate.month
    days = today.day - birthdate.day
        
    if days < 0:
        #get previous month and year
        previous_month = today.month - 1 if today.month > 1 else 12
        previous_year = today.year if today.month > 1 else today.year - 1
	#calculate days in previous month
        days_in_previous_month = (datetime(previous_year, previous_month ,1) + 
        			relativedelta(months=1) -                            
				datetime(previous_year, previous_month, 1)).days

        days += days_in_previous_month
        months -= 1
            
    if months < 0:
        years -= 1
        months += 12

    return years, months, days
    
def excel_table():
    excel_file_path = os.path.join(app.static_folder, 'PG_2024.xlsx')
    try:
        df = pd.read_excel(excel_file_path, sheet_name="points")
        
        if 'Unnamed: 0' in df.columns:
            df.rename(columns={'Unnamed: 0': 'Name'}, inplace=True)  
        
        df = df.loc[:, ~df.columns.str.contains('^Unnamed', na=False) | (df.columns == 'Name')]
        
        for col in df.columns:
            if col != 'Name':
                df[col] = pd.to_numeric(df[col], errors='coerce').astype('Int64')  
                df[col] = df[col].apply(lambda x: "" if pd.isna(x) else int(x))  

        table_html = df.to_html(classes='table-container', index=False)
        print("Generated HTML Table:", table_html)  
        return table_html
    except Exception as e:
            return f"Error loading Excel table: {str(e)}"
    

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/program')
def program():
    return render_template('program.html')

@app.route('/special', methods=['GET', 'POST'])
def special():
    if request.method == 'POST':
        try:
            name = request.form['name']
            
            if name in members_dict_datetime:
                birthdate = members_dict_datetime[name]
                session['result_2'] = calculate_indiv_age(birthdate)
                session['input_name'] = name
                session.pop('error_message', None)  # Clear any previous error

            else:
                session['error_message'] = "Name ungültig"
                session.pop('result_2', None)  # Clear any previous result
                session['input_name'] = name
                
        except Exception as e:
            session['error_message'] = f"An error occurred: {str(e)}"
            session.pop('result_2', None)  # Clear any previous result
        
        return redirect(url_for('special'))

    # Handle GET request
    if request.method == 'GET':
        result_1 = calculate_average_age()
        result_2 = session.get('result_2')
        error_message = session.get('error_message')
        input_name = session.get('input_name')
        result_3 = excel_table()

        # Clear session variables after rendering
        session.pop('result_2', None)
        session.pop('error_message', None)
        session.pop('input_name', None)

        return render_template(
            'special.html',
            result_1=result_1,
            result_2=result_2,
            result_3=result_3,
            error_message=error_message,
            input_name=input_name
        )

@app.route('/gallery')
def gallery():
    return render_template('gallery.html')

@app.after_request
def add_header(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '-1'
    return response

@app.context_processor
def utility_processor():
    return dict(version=str(time.time()))

if __name__ == '__main__':
    app.run(debug=True)
